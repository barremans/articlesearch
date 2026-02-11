# ======================================
# 🕒 ui_timings.py — V17.13 (FIXED)
# ======================================

import sys, os, re, shutil, tempfile, traceback, logging
from datetime import datetime, timedelta
import pandas as pd
import msal
from office365.sharepoint.client_context import ClientContext
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QListWidget,
    QListWidgetItem, QTextEdit, QProgressBar, QHBoxLayout,
    QCheckBox, QMessageBox, QLineEdit
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QKeySequence, QShortcut
from openpyxl import load_workbook

try:
    from permissions_azure import get_access_token
except ImportError:
    get_access_token = None


# ======================================
# 🪵 LOGGING
# ======================================
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

fh = logging.FileHandler("debug_log.txt", mode="w", encoding="utf-8")
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))

logger.addHandler(fh)
logger.addHandler(ch)

def safe_log(msg, level="info"):
    clean = re.sub(r"[^\x00-\x7F]+", "", str(msg))
    getattr(logging, level)(clean)


# ======================================
# 🔐 TOKEN MANAGER
# ======================================
class TokenManager:
    _token = None
    _expires_at = None

    TENANT = "526b32fa-8cb1-4d6a-9e2b-fd48e2a0e296"
    CLIENT = "58f55e10-e404-4307-9fa2-7b40431782fe"
    RESOURCE = "https://cgkgroupbvba.sharepoint.com/.default"

    @classmethod
    def get_token(cls, force_refresh=False):
        if cls._token and cls._expires_at and datetime.now() < cls._expires_at and not force_refresh:
            return cls._token

        try:
            if get_access_token:
                token = get_access_token(cls.RESOURCE)
                if token:
                    cls._token = token
                    cls._expires_at = datetime.now() + timedelta(hours=1)
                    return token
        except Exception:
            pass

        app = msal.PublicClientApplication(
            cls.CLIENT,
            authority=f"https://login.microsoftonline.com/{cls.TENANT}"
        )
        res = app.acquire_token_interactive(scopes=[cls.RESOURCE])
        cls._token = res["access_token"]
        cls._expires_at = datetime.now() + timedelta(seconds=res.get("expires_in", 3600))
        return cls._token


# ======================================
# 🧩 HULPFUNCTIES LOGICA (uit document 2)
# ======================================

def parse_excel_time(value):
    if pd.isna(value) or str(value).strip() == "":
        return None
    value = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            t = datetime.strptime(value, fmt)
            return datetime(2000, 1, 1, t.hour, t.minute, t.second)
        except ValueError:
            continue
    return None


def bereken_uren(row):
    try:
        start = parse_excel_time(row["Starttijd"])
        eind = parse_excel_time(row["Eindtijd"])
        if not start or not eind:
            return 0.0
        diff = (eind - start).total_seconds() / 3600
        if diff < 0:
            diff += 24
        return round(diff, 2)
    except Exception as e:
        safe_log(f"Fout bij bereken_uren: {e}", "debug")
        return 0.0


def schoon_data(df):
    df.columns = [c.strip() for c in df.columns]
    df = df.dropna(subset=["Naam", "Datum"], how="any")
    df["Naam"] = df["Naam"].astype(str).str.strip()
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce", dayfirst=True)

    for kol in ["Starttijd", "Eindtijd"]:
        if kol in df.columns:
            df[kol] = (
                df[kol]
                .astype(str)
                .str.replace("\xa0", "", regex=False)
                .str.replace(r"\s+", "", regex=True)
                .str.strip()
            )

    geldig_tijdpatroon = r"^\d{1,2}:\d{2}(:\d{2})?$"
    df = df[
        df["Starttijd"].str.match(geldig_tijdpatroon, na=False)
        & df["Eindtijd"].str.match(geldig_tijdpatroon, na=False)
    ]
    return df


def combineer_dubbels(df):
    kolommen_check = ["Datum", "Naam", "Starttijd", "Eindtijd"]
    if not all(k in df.columns for k in kolommen_check):
        safe_log("Kolommen voor samenvoegen ontbreken.", "warning")
        return df

    origineel_aantal = len(df)

    def tijd_in_minuten(t):
        t = parse_excel_time(t)
        if not t:
            return None
        return t.hour * 60 + t.minute

    df["Start_min"] = df["Starttijd"].apply(tijd_in_minuten)
    df["Eind_min"] = df["Eindtijd"].apply(tijd_in_minuten)
    df = df.dropna(subset=["Start_min", "Eind_min"])

    df = df.drop_duplicates(subset=["Naam", "Datum", "Start_min", "Eind_min"], keep="first")

    def filter_zero_uren(subdf):
        subdf["Uren_gewerkt"] = subdf.apply(bereken_uren, axis=1)
        if (subdf["Uren_gewerkt"] > 0).any():
            return subdf[subdf["Uren_gewerkt"] > 0]
        return subdf

    df = df.groupby(["Naam", "Datum"], group_keys=False).apply(filter_zero_uren)

    resultaten = []
    for (naam, datum), groep in df.groupby(["Naam", "Datum"]):
        tijden = sorted(zip(groep["Start_min"], groep["Eind_min"]))
        gecombineerd = []
        for start, eind in tijden:
            if not gecombineerd:
                gecombineerd.append([start, eind])
            else:
                laatste_start, laatste_eind = gecombineerd[-1]
                if start <= laatste_eind:
                    gecombineerd[-1][1] = max(laatste_eind, eind)
                else:
                    gecombineerd.append([start, eind])

        totaal_minuten = sum(e - s for s, e in gecombineerd)
        uren = round(totaal_minuten / 60, 2)
        intervallen_str = ", ".join(
            [f"{s//60:02.0f}:{s%60:02.0f}-{e//60:02.0f}:{e%60:02.0f}" for s, e in gecombineerd]
        )

        resultaten.append({
            "Naam": naam,
            "Datum": datum,
            "Uren_gewerkt": uren,
            "Eerste_start": f"{min(groep['Start_min'])//60:02.0f}:{min(groep['Start_min'])%60:02.0f}",
            "Laatste_eind": f"{max(groep['Eind_min'])//60:02.0f}:{max(groep['Eind_min'])%60:02.0f}",
            "Aantal_regels": len(groep),
            "Heeft_lege_regel": (groep["Uren_gewerkt"] == 0).any(),
            "Intervallen": intervallen_str
        })

    df_result = pd.DataFrame(resultaten)
    verwijderd = origineel_aantal - len(df_result)
    safe_log(f"Registraties gecombineerd: {origineel_aantal} → {len(df_result)} (verwijderd: {verwijderd})", "info")
    return df_result


def maak_overzicht(df):
    dag_map = {0: "Ma", 1: "Di", 2: "Wo", 3: "Do", 4: "Vr", 5: "Za", 6: "Zo"}
    df["Dag"] = df["Datum"].dt.dayofweek.map(dag_map)
    df_grouped = df.groupby(["Naam", "Dag"], as_index=False)["Uren_gewerkt"].sum()
    pivot = df_grouped.pivot(index="Naam", columns="Dag", values="Uren_gewerkt").fillna(0)
    vaste_dagen = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]
    for dag in vaste_dagen:
        if dag not in pivot.columns:
            pivot[dag] = 0
    pivot = pivot[vaste_dagen]
    pivot["Totaal"] = pivot.sum(axis=1).round(2)
    return pivot.reset_index()


def lees_timings_veilig(pad, xls):
    try:
        df = pd.read_excel(xls, "Timings")
        return df
    except Exception:
        wb = load_workbook(pad, data_only=True, read_only=True)
        ws = wb["Timings"]
        data = [[c or "" for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        return pd.DataFrame(data[1:], columns=data[0])


# ======================================
# 🔄 SHAREPOINT WORKER
# ======================================
class SharePointWorker(QThread):
    message = Signal(str)
    done = Signal(bool)
    files_list = Signal(list)
    progress = Signal(int)

    def __init__(self, mode="list", gekozen=None, auto_open=False):
        super().__init__()
        self.mode = mode
        self.gekozen = gekozen
        self.auto_open = auto_open

        self.download_dir = r"C:\temp"
        os.makedirs(self.download_dir, exist_ok=True)

        self.SITE = "https://cgkgroupbvba.sharepoint.com/sites/IT-team"
        self.PATH = "Gedeelde documenten/Applicaties/Evac_App/WeekLogs"

    def run(self):
        try:
            token = TokenManager.get_token()

            class Tok:
                def __init__(self, t):
                    self.tokenType = "Bearer"
                    self.accessToken = t

            ctx = ClientContext(self.SITE).with_access_token(lambda: Tok(token))
            folder = ctx.web.get_folder_by_server_relative_url(self.PATH)

            files = folder.files
            ctx.load(files)
            ctx.execute_query()

            items = []
            for f in files:
                naam = f.properties.get("Name", "")
                if not naam.lower().startswith("aanwezigheden_"):
                    continue

                lengte = float(f.properties.get("Length", 0)) / 1024
                raw_dt = f.properties.get("TimeLastModified")

                try:
                    dt = datetime.fromisoformat(str(raw_dt))
                    dstr = dt.strftime("%d-%m-%Y %H:%M")
                except Exception:
                    dt = datetime.min
                    dstr = "?"

                items.append({
                    "datum": dt,
                    "tekst": f"{naam} ({round(lengte,1)} KB, {dstr})"
                })

            # 🔥 sorteer: jongste eerst
            items.sort(key=lambda x: x["datum"], reverse=True)
            self.files_list.emit(items)

            if self.mode == "list":
                self.message.emit("📂 Lijst opgehaald — geen download.")
                self.done.emit(True)
                return

            clean = self.gekozen.split()[0]
            sel = next((f for f in files if f.properties["Name"] == clean), None)
            if not sel:
                self.message.emit("❌ Bestand niet gevonden.")
                self.done.emit(False)
                return

            path = os.path.join(self.download_dir, clean)
            with open(path, "wb") as fh:
                sel.download(fh).execute_query()

            self.message.emit(f"⬇️ Gedownload: {path}")

            # ✅ FIX 1: Voer verwerkingslogica uit
            self.message.emit("⚙️ Bestand verwerken...")
            self.verwerk_excel(path)
            self.message.emit("✅ Verwerking afgerond.")

            # ✅ FIX 2: Open bestand als auto_open=True
            if self.auto_open:
                self.message.emit("📂 Excel openen...")
                os.startfile(path)

            self.done.emit(True)

        except Exception as e:
            safe_log(traceback.format_exc(), "error")
            self.message.emit(f"❌ Fout: {e}")
            self.done.emit(False)

    def verwerk_excel(self, bestand):
        """
        ✅ TIMINGS-LOGICA UIT DOCUMENT 2
        Verwerk het Excel-bestand en sla het verwerkte resultaat op in C:\temp.
        """
        try:
            # Gebruik een tijdelijke kopie voor verwerking
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, "temp_uren.xlsx")
            shutil.copy(bestand, temp_path)
            xls = pd.ExcelFile(temp_path)

            # --- WeekData verwerken ---
            if "WeekData" in xls.sheet_names:
                df_week = pd.read_excel(xls, "WeekData")
                df_week = schoon_data(df_week)
                df_week["Uren_gewerkt"] = df_week.apply(bereken_uren, axis=1)
                overzicht_week = maak_overzicht(df_week)
                with pd.ExcelWriter(temp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    overzicht_week.to_excel(writer, sheet_name="WeekOverview", index=False)

            # --- Timings verwerken ---
            if "Timings" in xls.sheet_names:
                df_time = lees_timings_veilig(temp_path, xls)
                df_time = schoon_data(df_time)
                df_time["Uren_gewerkt"] = df_time.apply(bereken_uren, axis=1)
                df_time_samengevoegd = combineer_dubbels(df_time)
                overzicht_time = maak_overzicht(df_time_samengevoegd)
                debug_tab = df_time_samengevoegd.copy()
                debug_tab["Dag"] = debug_tab["Datum"].dt.strftime("%a")

                with pd.ExcelWriter(temp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    overzicht_time.to_excel(writer, sheet_name="WeekOverviewTimings", index=False)
                    debug_tab.to_excel(writer, sheet_name="TimingsOverview", index=False)

            # --- Verwerkt bestand opslaan in C:\temp ---
            output_path = os.path.join(self.download_dir, os.path.basename(bestand))
            shutil.copy2(temp_path, output_path)
            self.message.emit(f"💾 Verwerkt bestand opgeslagen: {output_path}")

        except Exception as e:
            tb = traceback.format_exc()
            safe_log(f"Fout tijdens verwerking: {e}\n{tb}", "error")
            self.message.emit(f"❌ Verwerkingsfout: {e}")


# ======================================
# 🖥️ GUI
# ======================================
class ExcelApp(QWidget):
    def __init__(self):
        super().__init__()

        self.setWindowTitle("Urenregistratie Downloader + Verwerker (V17.13)")
        self.setGeometry(400, 200, 850, 650)

        lay = QVBoxLayout(self)

        # 🔍 zoekveld
        self.search = QLineEdit()
        self.search.setPlaceholderText("🔎 Zoeken...")
        lay.addWidget(self.search)

        # ☑️ toon alles / top 5
        self.chk_all = QCheckBox("📂 Toon alle bestanden")
        self.chk_all.setChecked(False)   # standaard: top 5
        lay.addWidget(self.chk_all)

        # 📄 lijst
        self.list = QListWidget()
        lay.addWidget(self.list)

        # opties
        self.auto_open = QCheckBox("📂 Excel automatisch openen na verwerking")
        self.auto_open.setChecked(True)
        lay.addWidget(self.auto_open)

        # knoppen
        row = QHBoxLayout()
        self.btn_ref = QPushButton("🔄 Lijst verversen")
        self.btn_dl = QPushButton("⬇️ Download geselecteerd")
        row.addWidget(self.btn_ref)
        row.addWidget(self.btn_dl)
        lay.addLayout(row)

        # log
        self.log = QTextEdit()
        self.log.setReadOnly(True)
        lay.addWidget(self.log)

        self.all_items = []

        # events
        self.btn_ref.clicked.connect(self.laad)
        self.btn_dl.clicked.connect(self.start)
        self.search.textChanged.connect(self.filter)
        self.chk_all.stateChanged.connect(lambda _: self.filter(self.search.text()))

        QShortcut(QKeySequence(Qt.Key_Escape), self, activated=self.close)

        self.log.append("🔐 Authenticatie...")
        TokenManager.get_token()
        self.log.append("✅ Gereed.")

    # -----------------------------
    def laad(self):
        self.log.append("🌐 Ophalen lijst...")
        self.worker = SharePointWorker(mode="list")
        self.worker.files_list.connect(self.set_items)
        self.worker.message.connect(self.log.append)
        self.worker.start()

    def set_items(self, items):
        self.all_items = items
        self.filter("")

    def filter(self, txt):
        self.list.clear()
        t = txt.lower()

        items = self.all_items
        if not self.chk_all.isChecked():
            items = items[:5]   # ⭐ standaard top 5

        for i in items:
            if t in i["tekst"].lower():
                self.list.addItem(QListWidgetItem(i["tekst"]))

    def start(self):
        it = self.list.currentItem()
        if not it:
            QMessageBox.warning(self, "Geen selectie", "Selecteer een bestand.")
            return

        self.worker = SharePointWorker(
            mode="download",
            gekozen=it.text(),
            auto_open=self.auto_open.isChecked()
        )
        self.worker.message.connect(self.log.append)
        self.worker.start()


# ======================================
# 🚀 MAIN
# ======================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = ExcelApp()
    w.show()
    sys.exit(app.exec())