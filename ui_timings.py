#!/usr/bin/env python3
# ======================================
# 🕒 ui_timings.py — VERBETERDE VERSIE
#V17.15
# Timing Logica: Verwerkt IN/OUT badges correct
# ======================================

import sys
import os
import re
import shutil
import tempfile
import traceback
import logging
from datetime import datetime, timedelta
import pandas as pd
import msal
from office365.sharepoint.client_context import ClientContext
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QListWidget,
    QListWidgetItem, QTextEdit, QHBoxLayout,
    QCheckBox, QMessageBox, QLineEdit
)
from PySide6.QtCore import Qt, QThread, Signal
from PySide6.QtGui import QKeySequence, QShortcut
from openpyxl import load_workbook

try:
    from permissions_azure import get_access_token
except ImportError:
    get_access_token = None

# ======================================
# 🪵 LOGGING
# ======================================
logger = logging.getLogger()
logger.setLevel(logging.DEBUG)

fh = logging.FileHandler("debug_log.txt", mode="w", encoding="utf-8")
fh.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))

ch = logging.StreamHandler(sys.stdout)
ch.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))

logger.addHandler(fh)
logger.addHandler(ch)


def safe_log(msg, level="info"):
    """Log message zonder speciale karakters"""
    clean = re.sub(r"[^\x00-\x7F]+", "", str(msg))
    getattr(logging, level)(clean)


# ======================================
# 🔐 TOKEN MANAGER
# ======================================
class TokenManager:
    """Beheer SharePoint authenticatie tokens"""
    _token = None
    _expires_at = None

    TENANT = "526b32fa-8cb1-4d6a-9e2b-fd48e2a0e296"
    CLIENT = "58f55e10-e404-4307-9fa2-7b40431782fe"
    RESOURCE = "https://cgkgroupbvba.sharepoint.com/.default"

    @classmethod
    def get_token(cls):
        """Haal een geldig access token op"""
        if cls._token and cls._expires_at and datetime.now() < cls._expires_at:
            return cls._token

        # Probeer silent token
        if get_access_token:
            try:
                token = get_access_token(cls.RESOURCE)
                if token:
                    cls._token = token
                    cls._expires_at = datetime.now() + timedelta(hours=1)
                    return token
            except Exception:
                pass

        # Fallback naar interactive login
        app = msal.PublicClientApplication(
            cls.CLIENT,
            authority=f"https://login.microsoftonline.com/{cls.TENANT}"
        )
        res = app.acquire_token_interactive(scopes=[cls.RESOURCE])
        cls._token = res["access_token"]
        cls._expires_at = datetime.now() + timedelta(seconds=res.get("expires_in", 3600))
        return cls._token


# ======================================
# 🧩 TIMING LOGICA (VERBETERD)
# ======================================

def parse_time(value):
    """Parse tijd naar datetime object"""
    if pd.isna(value):
        return None
    
    s = str(value).strip()
    
    # Probeer verschillende formaten
    for fmt in ["%H:%M:%S", "%H:%M"]:
        try:
            t = datetime.strptime(s, fmt)
            return datetime(2000, 1, 1, t.hour, t.minute, t.second)
        except ValueError:
            continue
    
    return None


def bereken_uren(start, eind):
    """Bereken uren tussen twee tijden"""
    if not start or not eind:
        return 0.0
    
    diff = (eind - start).total_seconds() / 3600
    if diff < 0:
        diff += 24  # Voor shifts over middernacht
    
    return round(diff, 2)


# Dag mapping
DAG_MAP = {
    "Ma": "Ma", "Di": "Di", "Wo": "Wo", "Do": "Do",
    "Vr": "Vr", "Za": "Za", "Zo": "Zo",
    "Mon": "Ma", "Tue": "Di", "Wed": "Wo",
    "Thu": "Do", "Fri": "Vr", "Sat": "Za", "Sun": "Zo",
    "Monday": "Ma", "Tuesday": "Di", "Wednesday": "Wo",
    "Thursday": "Do", "Friday": "Vr", "Saturday": "Za", "Sunday": "Zo"
}


def schoon_data(df):
    """Maak data schoon en verwijder ongeldige entries"""
    df = df.copy()
    df.columns = [c.strip() for c in df.columns]

    df = df.dropna(subset=["Naam", "Datum"], how="any")
    df["Naam"] = df["Naam"].astype(str).str.strip()
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce", dayfirst=True)

    if "Dag" in df.columns:
        df["Dag"] = (
            df["Dag"]
            .astype(str)
            .str.strip()
            .map(DAG_MAP)
        )

    for kol in ["Starttijd", "Eindtijd"]:
        if kol in df.columns:
            df[kol] = (
                df[kol].astype(str)
                .str.replace("\xa0", "", regex=False)
                .str.replace(r"\s+", "", regex=True)
                .str.strip()
            )

    geldig = r"^\d{1,2}:\d{2}(:\d{2})?$"
    df = df[
        df["Starttijd"].str.match(geldig, na=False)
        & df["Eindtijd"].str.match(geldig, na=False)
    ]
    return df


def verwerk_timings(df):
    """
    Verwerk timings en identificeer IN/OUT badges correct
    
    IN badge: starttijd == eindtijd (bijv. 06:55:00 - 06:55:00)
    OUT badge: starttijd != eindtijd (bijv. 06:55:00 - 15:49:00)
    """
    
    # Parse tijden
    df['Start_parsed'] = df['Starttijd'].apply(parse_time)
    df['Eind_parsed'] = df['Eindtijd'].apply(parse_time)
    
    # Identificeer IN vs OUT badges
    # IN badge: starttijd == eindtijd (of zeer klein verschil < 1 minuut)
    df['Is_IN'] = df.apply(lambda row: 
        abs((row['Eind_parsed'] - row['Start_parsed']).total_seconds()) < 60 
        if row['Start_parsed'] and row['Eind_parsed'] else False, 
        axis=1
    )
    
    # Maak detail overzicht per persoon per dag
    detail_data = []
    overzicht_data = []
    
    for (naam, datum, dag), groep in df.groupby(['Naam', 'Datum', 'Dag']):
        # Sorteer op starttijd
        groep = groep.sort_values('Start_parsed')
        
        # Haal locatie op (eerste niet-NaN)
        locatie = groep['Locatie'].dropna().iloc[0] if len(groep['Locatie'].dropna()) > 0 else 'Onbekend'
        
        # Verzamel alle IN en OUT tijden
        in_tijden = []
        out_tijden = []
        
        for _, row in groep.iterrows():
            if row['Is_IN']:
                # Dit is een IN badge
                tijd_str = row['Start_parsed'].strftime('%H:%M')
                in_tijden.append(tijd_str)
            else:
                # Dit is een OUT badge (met berekende eindtijd)
                start_str = row['Start_parsed'].strftime('%H:%M')
                eind_str = row['Eind_parsed'].strftime('%H:%M')
                
                # Voeg IN toe als die nog niet bestaat
                if start_str not in in_tijden:
                    in_tijden.append(start_str)
                
                out_tijden.append(eind_str)
        
        # Combineer IN en OUT tijden tot periodes
        periodes = []
        totaal_uren = 0.0
        
        # Match IN tijden met OUT tijden
        for i, in_tijd in enumerate(in_tijden):
            in_dt = datetime.strptime(in_tijd, '%H:%M')
            
            # Zoek bijbehorende OUT tijd
            if i < len(out_tijden):
                out_tijd = out_tijden[i]
                out_dt = datetime.strptime(out_tijd, '%H:%M')
                
                # Bereken uren
                uren = bereken_uren(in_dt, out_dt)
                totaal_uren += uren
                
                periodes.append(f"{in_tijd}-{out_tijd} ({uren}u)")
            else:
                # Geen OUT badge (nog aan het werk?)
                periodes.append(f"{in_tijd}-? (actief)")
        
        # Detail entry
        detail_data.append({
            'Datum': datum.strftime('%d-%m-%Y'),
            'Dag': dag,
            'Naam': naam,
            'Locatie': locatie,
            'IN_badges': ', '.join(in_tijden) if in_tijden else '-',
            'OUT_badges': ', '.join(out_tijden) if out_tijden else '-',
            'Periodes': ' | '.join(periodes) if periodes else '-',
            'Totaal_uren': round(totaal_uren, 2)
        })
        
        # Overzicht entry (voor pivot)
        overzicht_data.append({
            'Naam': naam,
            'Datum': datum,
            'Dag': dag,
            'Locatie': locatie,
            'Uren': round(totaal_uren, 2)
        })
    
    return pd.DataFrame(detail_data), pd.DataFrame(overzicht_data)


def maak_week_overzicht(df_overzicht):
    """Maak een week overzicht met totalen per dag"""
    dagen_orde = ['Ma', 'Di', 'Wo', 'Do', 'Vr', 'Za', 'Zo']
    
    if df_overzicht.empty:
        return pd.DataFrame(columns=['Naam'] + dagen_orde + ['Totaal'])
    
    pivot = df_overzicht.pivot_table(
        index='Naam',
        columns='Dag',
        values='Uren',
        aggfunc='sum',
        fill_value=0
    )
    
    # Zorg dat alle dagen aanwezig zijn
    for dag in dagen_orde:
        if dag not in pivot.columns:
            pivot[dag] = 0
    
    pivot = pivot[dagen_orde]
    pivot['Totaal'] = pivot.sum(axis=1).round(2)
    pivot = pivot.reset_index()
    
    return pivot


def lees_timings_veilig(pad, xls):
    """Lees Timings sheet op een veilige manier"""
    try:
        return pd.read_excel(xls, "Timings")
    except Exception:
        wb = load_workbook(pad, data_only=True, read_only=True)
        ws = wb["Timings"]
        data = [[c or "" for c in r] for r in ws.iter_rows(values_only=True)]
        wb.close()
        return pd.DataFrame(data[1:], columns=data[0])


# ======================================
# 🔄 SHAREPOINT WORKER
# ======================================
class SharePointWorker(QThread):
    """Background thread voor SharePoint operaties"""
    message = Signal(str)
    files_list = Signal(list)
    done = Signal(bool)

    SITE = "https://cgkgroupbvba.sharepoint.com/sites/IT-team"
    PATH = "Gedeelde documenten/Applicaties/Evac_App/WeekLogs"

    def __init__(self, mode="list", gekozen=None, auto_open=False):
        super().__init__()
        self.mode = mode
        self.gekozen = gekozen
        self.auto_open = auto_open
        self.download_dir = r"C:\temp"
        os.makedirs(self.download_dir, exist_ok=True)

    def run(self):
        try:
            token = TokenManager.get_token()

            class Tok:
                def __init__(self, t):
                    self.tokenType = "Bearer"
                    self.accessToken = t

            ctx = ClientContext(self.SITE).with_access_token(lambda: Tok(token))
            folder = ctx.web.get_folder_by_server_relative_url(self.PATH)
            files = folder.files
            ctx.load(files)
            ctx.execute_query()

            items = []
            for f in files:
                name = f.properties["Name"]
                if not name.lower().startswith("aanwezigheden_"):
                    continue
                dt = datetime.fromisoformat(str(f.properties.get("TimeLastModified")))
                size = round(float(f.properties.get("Length", 0)) / 1024, 1)
                items.append({"dt": dt, "txt": f"{name} ({size} KB, {dt:%d-%m-%Y %H:%M})"})

            items.sort(key=lambda x: x["dt"], reverse=True)
            self.files_list.emit(items)

            if self.mode == "list":
                self.done.emit(True)
                return

            clean = self.gekozen.split()[0]
            sel = next(f for f in files if f.properties["Name"] == clean)

            path = os.path.join(self.download_dir, clean)
            with open(path, "wb") as fh:
                sel.download(fh).execute_query()

            self.verwerk_excel(path)

            if self.auto_open:
                os.startfile(path)

            self.done.emit(True)

        except Exception as e:
            safe_log(traceback.format_exc(), "error")
            self.message.emit(f"❌ Fout: {e}")
            self.done.emit(False)

    def verwerk_excel(self, bestand):
        """Verwerk Excel bestand met verbeterde timing logica"""
        tmp = os.path.join(tempfile.gettempdir(), "temp_uren.xlsx")
        shutil.copy(bestand, tmp)
        xls = pd.ExcelFile(tmp)

        with pd.ExcelWriter(tmp, engine="openpyxl", mode="a", if_sheet_exists="replace") as w:
            # Verwerk WeekData (indien aanwezig)
            if "WeekData" in xls.sheet_names:
                df = schoon_data(pd.read_excel(xls, "WeekData"))
                if not df.empty:
                    df_detail, df_overzicht = verwerk_timings(df)
                    df_week = maak_week_overzicht(df_overzicht)
                    
                    df_detail.to_excel(w, "WeekData_Detail", index=False)
                    df_week.to_excel(w, "WeekOverview", index=False)

            # Verwerk Timings (belangrijkste sheet)
            if "Timings" in xls.sheet_names:
                df = schoon_data(lees_timings_veilig(tmp, xls))
                if not df.empty:
                    df_detail, df_overzicht = verwerk_timings(df)
                    df_week = maak_week_overzicht(df_overzicht)
                    
                    df_detail.to_excel(w, "Timings_Detail", index=False)
                    df_week.to_excel(w, "WeekOverviewTimings", index=False)

        shutil.copy2(tmp, bestand)
        safe_log(f"✅ Excel verwerkt: {bestand}")


# ======================================
# 🖥️ GUI
# ======================================
class ExcelApp(QWidget):
    """Hoofd GUI applicatie"""
    
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Urenregistratie Downloader + Verwerker (VERBETERD)")
        self.setGeometry(400, 200, 850, 650)

        lay = QVBoxLayout(self)
        self.search = QLineEdit(placeholderText="🔎 Zoeken...")
        self.chk_all = QCheckBox("📂 Toon alle bestanden")
        self.list = QListWidget()
        self.auto_open = QCheckBox("📂 Excel automatisch openen", checked=True)
        self.log = QTextEdit(readOnly=True)

        btns = QHBoxLayout()
        self.btn_ref = QPushButton("🔄 Verversen")
        self.btn_dl = QPushButton("⬇️ Downloaden & Verwerken")
        btns.addWidget(self.btn_ref)
        btns.addWidget(self.btn_dl)

        for w in [self.search, self.chk_all, self.list, self.auto_open, btns, self.log]:
            lay.addWidget(w) if not isinstance(w, QHBoxLayout) else lay.addLayout(w)

        self.items = []
        self.btn_ref.clicked.connect(self.laad)
        self.btn_dl.clicked.connect(self.start)
        self.search.textChanged.connect(self.filter)
        self.chk_all.stateChanged.connect(lambda _: self.filter(self.search.text()))
        QShortcut(QKeySequence(Qt.Key_Escape), self, activated=self.close)

        # Pre-load token
        try:
            TokenManager.get_token()
            self.log.append("✅ Authenticatie succesvol")
        except Exception as e:
            self.log.append(f"⚠️ Authenticatie fout: {e}")

    def laad(self):
        """Laad lijst met bestanden van SharePoint"""
        self.log.append("📥 Bestanden ophalen...")
        self.worker = SharePointWorker(mode="list")
        self.worker.files_list.connect(self.set_items)
        self.worker.message.connect(self.log.append)
        self.worker.done.connect(lambda ok: self.log.append("✅ Lijst geladen" if ok else "❌ Fout bij laden"))
        self.worker.start()

    def set_items(self, items):
        """Zet items in de lijst"""
        self.items = items
        self.filter("")

    def filter(self, txt):
        """Filter lijst op zoektekst"""
        self.list.clear()
        items = self.items if self.chk_all.isChecked() else self.items[:5]
        for i in items:
            if txt.lower() in i["txt"].lower():
                self.list.addItem(QListWidgetItem(i["txt"]))

    def start(self):
        """Start download en verwerking"""
        it = self.list.currentItem()
        if not it:
            QMessageBox.warning(self, "Geen selectie", "Selecteer een bestand.")
            return
        
        self.log.append(f"⬇️ Downloaden: {it.text()}")
        self.worker = SharePointWorker(
            mode="download",
            gekozen=it.text(),
            auto_open=self.auto_open.isChecked()
        )
        self.worker.message.connect(self.log.append)
        self.worker.done.connect(lambda ok: self.log.append(
            "✅ Verwerking compleet!" if ok else "❌ Fout bij verwerking"
        ))
        self.worker.start()


# ======================================
# 🚀 MAIN
# ======================================
if __name__ == "__main__":
    app = QApplication(sys.argv)
    w = ExcelApp()
    w.show()
    sys.exit(app.exec())