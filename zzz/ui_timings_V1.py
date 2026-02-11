# ======================================
# 🕒 ui_timings.py — Urenregistratie Verwerker + SharePoint download
# ======================================

import sys
import os
import re
import shutil
import tempfile
import traceback
import logging
from datetime import datetime, timedelta
import pandas as pd
import msal
from office365.sharepoint.client_context import ClientContext
from PySide6.QtWidgets import (
    QApplication, QWidget, QVBoxLayout, QLabel, QPushButton, QLineEdit,
    QMessageBox, QTextEdit, QProgressBar
)
from PySide6.QtCore import Qt, QThread, Signal
from openpyxl import load_workbook

from permissions_azure import get_access_token


# ======================================
# 🪵 LOGGING
# ======================================

logger = logging.getLogger()
logger.setLevel(logging.DEBUG)
file_handler = logging.FileHandler("debug_log.txt", mode="w", encoding="utf-8")
file_handler.setFormatter(logging.Formatter("%(asctime)s [%(levelname)s] %(message)s"))
console_handler = logging.StreamHandler(sys.stdout)
console_handler.setFormatter(logging.Formatter("[%(levelname)s] %(message)s"))
logger.addHandler(file_handler)
logger.addHandler(console_handler)


def safe_log(msg: str, level="info"):
    clean = re.sub(r"[^\x00-\x7F]+", "", msg)
    getattr(logging, level)(clean)


# ======================================
# 🧩 HULPFUNCTIES LOGICA
# ======================================

def parse_excel_time(value):
    if pd.isna(value) or str(value).strip() == "":
        return None
    value = str(value).strip()
    for fmt in ("%H:%M:%S", "%H:%M"):
        try:
            t = datetime.strptime(value, fmt)
            return datetime(2000, 1, 1, t.hour, t.minute, t.second)
        except ValueError:
            continue
    return None


def bereken_uren(row):
    try:
        start = parse_excel_time(row["Starttijd"])
        eind = parse_excel_time(row["Eindtijd"])
        if not start or not eind:
            return 0.0
        diff = (eind - start).total_seconds() / 3600
        if diff < 0:
            diff += 24
        return round(diff, 2)
    except Exception as e:
        safe_log(f"Fout bij bereken_uren: {e}", "debug")
        return 0.0


def schoon_data(df):
    df.columns = [c.strip() for c in df.columns]
    df = df.dropna(subset=["Naam", "Datum"], how="any")
    df["Naam"] = df["Naam"].astype(str).str.strip()
    df["Datum"] = pd.to_datetime(df["Datum"], errors="coerce", dayfirst=True)

    for kol in ["Starttijd", "Eindtijd"]:
        if kol in df.columns:
            df[kol] = (
                df[kol]
                .astype(str)
                .str.replace("\xa0", "", regex=False)
                .str.replace(r"\s+", "", regex=True)
                .str.strip()
            )

    geldig_tijdpatroon = r"^\d{1,2}:\d{2}(:\d{2})?$"
    df = df[
        df["Starttijd"].str.match(geldig_tijdpatroon, na=False)
        & df["Eindtijd"].str.match(geldig_tijdpatroon, na=False)
    ]
    return df


def combineer_dubbels(df):
    kolommen_check = ["Datum", "Naam", "Starttijd", "Eindtijd"]
    if not all(k in df.columns for k in kolommen_check):
        safe_log("Kolommen voor samenvoegen ontbreken.", "warning")
        return df

    origineel_aantal = len(df)

    def tijd_in_minuten(t):
        t = parse_excel_time(t)
        if not t:
            return None
        return t.hour * 60 + t.minute

    df["Start_min"] = df["Starttijd"].apply(tijd_in_minuten)
    df["Eind_min"] = df["Eindtijd"].apply(tijd_in_minuten)
    df = df.dropna(subset=["Start_min", "Eind_min"])

    df = df.drop_duplicates(subset=["Naam", "Datum", "Start_min", "Eind_min"], keep="first")

    def filter_zero_uren(subdf):
        subdf["Uren_gewerkt"] = subdf.apply(bereken_uren, axis=1)
        if (subdf["Uren_gewerkt"] > 0).any():
            return subdf[subdf["Uren_gewerkt"] > 0]
        return subdf

    df = df.groupby(["Naam", "Datum"], group_keys=False).apply(filter_zero_uren)

    resultaten = []
    for (naam, datum), groep in df.groupby(["Naam", "Datum"]):
        tijden = sorted(zip(groep["Start_min"], groep["Eind_min"]))
        gecombineerd = []
        for start, eind in tijden:
            if not gecombineerd:
                gecombineerd.append([start, eind])
            else:
                laatste_start, laatste_eind = gecombineerd[-1]
                if start <= laatste_eind:
                    gecombineerd[-1][1] = max(laatste_eind, eind)
                else:
                    gecombineerd.append([start, eind])

        totaal_minuten = sum(e - s for s, e in gecombineerd)
        uren = round(totaal_minuten / 60, 2)
        intervallen_str = ", ".join(
            [f"{s//60:02.0f}:{s%60:02.0f}-{e//60:02.0f}:{e%60:02.0f}" for s, e in gecombineerd]
        )

        resultaten.append({
            "Naam": naam,
            "Datum": datum,
            "Uren_gewerkt": uren,
            "Eerste_start": f"{min(groep['Start_min'])//60:02.0f}:{min(groep['Start_min'])%60:02.0f}",
            "Laatste_eind": f"{max(groep['Eind_min'])//60:02.0f}:{max(groep['Eind_min'])%60:02.0f}",
            "Aantal_regels": len(groep),
            "Heeft_lege_regel": (groep["Uren_gewerkt"] == 0).any(),
            "Intervallen": intervallen_str
        })

    df_result = pd.DataFrame(resultaten)
    verwijderd = origineel_aantal - len(df_result)
    safe_log(f"Registraties gecombineerd: {origineel_aantal} → {len(df_result)} (verwijderd: {verwijderd})", "info")
    return df_result


def maak_overzicht(df):
    dag_map = {0: "Ma", 1: "Di", 2: "Wo", 3: "Do", 4: "Vr", 5: "Za", 6: "Zo"}
    df["Dag"] = df["Datum"].dt.dayofweek.map(dag_map)
    df_grouped = df.groupby(["Naam", "Dag"], as_index=False)["Uren_gewerkt"].sum()
    pivot = df_grouped.pivot(index="Naam", columns="Dag", values="Uren_gewerkt").fillna(0)
    vaste_dagen = ["Ma", "Di", "Wo", "Do", "Vr", "Za", "Zo"]
    for dag in vaste_dagen:
        if dag not in pivot.columns:
            pivot[dag] = 0
    pivot = pivot[vaste_dagen]
    pivot["Totaal"] = pivot.sum(axis=1).round(2)
    return pivot.reset_index()


def lees_timings_veilig(pad, xls):
    try:
        df = pd.read_excel(xls, "Timings")
        return df
    except Exception:
        wb = load_workbook(pad, data_only=True, read_only=True)
        ws = wb["Timings"]
        data = [[c or "" for c in row] for row in ws.iter_rows(values_only=True)]
        wb.close()
        return pd.DataFrame(data[1:], columns=data[0])


# ======================================
# 🔄 SHAREPOINT DOWNLOADER + VERWERKER
# ======================================

class SharePointWorker(QThread):
    message = Signal(str)
    done = Signal(bool)

    def __init__(self, simulatie_datum=None):
        super().__init__()
        self.simulatiedatum = simulatie_datum
        # ✅ Downloadmap: altijd C:\temp
        self.download_path = r"C:\temp"
        if not os.path.exists(self.download_path):
            try:
                os.makedirs(self.download_path, exist_ok=True)
                print(f"[INFO] Downloadmap aangemaakt: {self.download_path}")
            except Exception as e:
                print(f"[WARN] Kon {self.download_path} niet aanmaken: {e}")
                # fallback naar lokale temp als C:\temp niet beschikbaar is
                self.download_path = os.path.join(os.getcwd(), "temp")
                os.makedirs(self.download_path, exist_ok=True)


        self.TENANT_ID = "526b32fa-8cb1-4d6a-9e2b-fd48e2a0e296"
        self.CLIENT_ID = "58f55e10-e404-4307-9fa2-7b40431782fe"
        self.SHAREPOINT_SITE = "https://cgkgroupbvba.sharepoint.com/sites/IT-team"
        self.LIBRARY_PATH = "Gedeelde documenten/Applicaties/Evac_App/WeekLogs"

    def run(self):
        try:
            self.message.emit("🌐 Verbinden met Microsoft 365...")

            # ✅ Token ophalen via permissions_azure (hergebruikt bestaande login)
            access_token = get_access_token("https://cgkgroupbvba.sharepoint.com/.default")

            # Token wrapper voor Office365 library
            class TokenResponse:
                def __init__(self, access_token):
                    self.tokenType = "Bearer"
                    self.accessToken = access_token

            token = TokenResponse(access_token)
            ctx = ClientContext(self.SHAREPOINT_SITE).with_access_token(lambda: token)

            # 🔍 Bepaal doelbestand (maandag van deze of simulatie week)
            gekozen_datum = (
                datetime.strptime(self.simulatiedatum, "%d-%m-%Y")
                if self.simulatiedatum
                else datetime.today()
            )
            maandag = gekozen_datum - timedelta(days=gekozen_datum.weekday())
            target = f"Aanwezigheden_{maandag.strftime('%d-%m-%Y')}.xlsx"

            self.message.emit(f"📁 Zoeken naar bestand: {target}")
            folder = ctx.web.get_folder_by_server_relative_url(self.LIBRARY_PATH)
            files = folder.files
            ctx.load(files)
            ctx.execute_query()

            gekozen_file = next((f for f in files if f.properties["Name"].lower() == target.lower()), None)
            if not gekozen_file:
                self.message.emit(f"❌ Bestand '{target}' niet gevonden.")
                self.done.emit(False)
                return

            local_path = os.path.join(self.download_path, target)
            with open(local_path, "wb") as f:
                gekozen_file.download(f).execute_query()
            self.message.emit(f"✅ Bestand gedownload: {local_path}")

            # ⚙️ Verwerken
            self.message.emit("⚙️ Bestand verwerken...")
            self.verwerk_excel(local_path)
            self.message.emit("✅ Verwerking afgerond. Bestand openen...")
            os.startfile(local_path)
            self.done.emit(True)

        except Exception as e:
            tb = traceback.format_exc()
            safe_log(f"SharePoint fout: {e}\n{tb}", "error")
            self.message.emit(f"❌ Fout: {e}")
            self.done.emit(False)


    def verwerk_excel(self, bestand):
        """
        Verwerk het Excel-bestand en sla het verwerkte resultaat op in C:\temp.
        """
        try:
            # Gebruik een tijdelijke kopie voor verwerking
            temp_dir = tempfile.gettempdir()
            temp_path = os.path.join(temp_dir, "temp_uren.xlsx")
            shutil.copy(bestand, temp_path)
            xls = pd.ExcelFile(temp_path)

            # --- WeekData verwerken ---
            if "WeekData" in xls.sheet_names:
                df_week = pd.read_excel(xls, "WeekData")
                df_week = schoon_data(df_week)
                df_week["Uren_gewerkt"] = df_week.apply(bereken_uren, axis=1)
                overzicht_week = maak_overzicht(df_week)
                with pd.ExcelWriter(temp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    overzicht_week.to_excel(writer, sheet_name="WeekOverview", index=False)

            # --- Timings verwerken ---
            if "Timings" in xls.sheet_names:
                df_time = lees_timings_veilig(temp_path, xls)
                df_time = schoon_data(df_time)
                df_time["Uren_gewerkt"] = df_time.apply(bereken_uren, axis=1)
                df_time_samengevoegd = combineer_dubbels(df_time)
                overzicht_time = maak_overzicht(df_time_samengevoegd)
                debug_tab = df_time_samengevoegd.copy()
                debug_tab["Dag"] = debug_tab["Datum"].dt.strftime("%a")

                with pd.ExcelWriter(temp_path, engine="openpyxl", mode="a", if_sheet_exists="replace") as writer:
                    overzicht_time.to_excel(writer, sheet_name="WeekOverviewTimings", index=False)
                    debug_tab.to_excel(writer, sheet_name="TimingsOverview", index=False)

            # --- Verwerkt bestand opslaan in C:\temp ---
            output_path = os.path.join(self.download_path, os.path.basename(bestand))
            shutil.copy2(temp_path, output_path)
            self.message.emit(f"📂 Verwerkt bestand opgeslagen in: {output_path}")

        except Exception as e:
            tb = traceback.format_exc()
            safe_log(f"Fout tijdens verwerking: {e}\n{tb}", "error")
            self.message.emit(f"❌ Verwerkingsfout: {e}")


# ======================================
# 🖥️ GUI
# ======================================

class ExcelApp(QWidget):
    def __init__(self):
        super().__init__()
        self.setWindowTitle("Urenregistratie Downloader + Verwerker")
        from PySide6.QtGui import QIcon
        self.setWindowIcon(QIcon(os.path.join(os.path.dirname(__file__), "assets", "clock.png")))

        self.setGeometry(400, 200, 700, 500)

        layout = QVBoxLayout()
        self.setLayout(layout)

        layout.addWidget(QLabel("Simulatiedatum (dd-mm-jjjj) - laat leeg voor deze week:"))
        self.sim_input = QLineEdit()
        layout.addWidget(self.sim_input)

        self.btn = QPushButton("⬇️ Download & Verwerk")
        self.btn.clicked.connect(self.run_worker)
        layout.addWidget(self.btn)

        self.progress = QProgressBar()
        self.progress.setValue(0)
        layout.addWidget(self.progress)

        self.log = QTextEdit()
        self.log.setReadOnly(True)
        layout.addWidget(self.log)

        self.worker = None
        
        # --- Sneltoetsen ---
        from PySide6.QtGui import QKeySequence, QShortcut
        QShortcut(QKeySequence(Qt.Key_Escape), self).activated.connect(self.close)
        QShortcut(QKeySequence(Qt.Key_Delete), self).activated.connect(self.clear_input)
        QShortcut(QKeySequence("Ctrl+D"), self).activated.connect(self.clear_input)
        QShortcut(QKeySequence("Ctrl+Return"), self).activated.connect(self.run_worker)

        
    def clear_input(self):
        """Leeg het invoerveld en zet de focus erop."""
        self.sim_input.clear()
        self.sim_input.setFocus()



    def append_log(self, text):
        self.log.append(text)
        self.log.verticalScrollBar().setValue(self.log.verticalScrollBar().maximum())

    def run_worker(self):
        sim_date = self.sim_input.text().strip()
        self.append_log(f"▶️ Start proces ({sim_date or 'huidige week'})...")
        self.worker = SharePointWorker(sim_date if sim_date else None)
        self.worker.message.connect(self.append_log)
        self.worker.done.connect(self.done)
        self.worker.start()

    def done(self, success):
        if success:
            QMessageBox.information(self, "Klaar", "✅ Download & verwerking afgerond.")
        else:
            QMessageBox.critical(self, "Fout", "❌ Er trad een fout op tijdens het proces.")


# ======================================
# 🚀 MAIN
# ======================================

if __name__ == "__main__":
    app = QApplication(sys.argv)
    window = ExcelApp()
    window.show()
    sys.exit(app.exec())
